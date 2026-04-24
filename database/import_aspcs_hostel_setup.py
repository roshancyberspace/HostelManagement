from __future__ import annotations

import sqlite3
from collections import defaultdict
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook


BASE_DIR = Path(__file__).resolve().parent
DB_PATH = BASE_DIR / "hostel.db"
STUDENTS_XLSX = Path(r"C:\Users\jis_3\Downloads\ASPCS HOSTEL STUDENTS LIST 2026-27.xlsx")

HOSTEL_NAME = "Sudarshan Ashram"
BLOCK_NAME = "Main Block"
TODAY = date.today().isoformat()
NOW = datetime.now().strftime("%Y-%m-%d %H:%M:%S")


@dataclass
class RoomDef:
    room_no: str
    room_type: str
    ac_type: str = "NON_AC"
    bed_capacity: int = 0
    status: str = "ACTIVE"
    is_merged: int = 0
    merged_into_room_no: str | None = None
    items: list[tuple[str, int]] | None = None


def room(room_no: str, room_type: str, ac_type: str = "NON_AC", bed_capacity: int = 0,
         status: str = "ACTIVE", is_merged: int = 0, merged_into_room_no: str | None = None,
         items: list[tuple[str, int]] | None = None) -> RoomDef:
    return RoomDef(
        room_no=room_no,
        room_type=room_type,
        ac_type=ac_type,
        bed_capacity=bed_capacity,
        status=status,
        is_merged=is_merged,
        merged_into_room_no=merged_into_room_no,
        items=items or [],
    )


FLOOR_COMMON_AREAS: dict[str, list[tuple[str, list[tuple[str, int]], str]]] = {
    "Ground Floor": [
        ("Reception Area", [("Table", 1), ("Chair", 3)], "GOOD"),
        ("Visitor Lobby", [("Sofa 3-Seater", 2)], "GOOD"),
        ("Account / Mess Manager Office", [("Table", 2), ("Chair", 3), ("Almirah", 1)], "GOOD"),
        ("G1 Bathroom", [("Latrin", 3), ("Bathroom Area", 3), ("Gyser", 1), ("Urinal", 1), ("Washbasin", 1)], "GOOD"),
        ("G2 Bathroom", [("Latrin", 3), ("Bathroom Area", 3), ("Gyser", 1), ("Urinal", 1)], "GOOD"),
        ("Dining Area", [], "GOOD"),
        ("Kitchen", [], "GOOD"),
        ("Kitchen Store", [], "GOOD"),
        ("Badminton Court", [], "GOOD"),
        ("Open Court", [], "GOOD"),
    ],
    "First Floor": [
        ("F1 Bathroom", [("Gyser", 1), ("Washbasin", 1), ("Latrin", 4), ("Bath Area", 4), ("Urinal", 1)], "GOOD"),
        ("F2 Bathroom", [("Damaged Gyser", 1), ("Latrin", 4), ("Bath Area", 4)], "GOOD"),
        ("F3 Bathroom", [("Gyser", 1), ("Latrin", 4), ("Wash Area", 4)], "GOOD"),
        ("F4 Bathroom", [("Gyser", 1), ("Urinal", 3), ("Latrin", 4), ("Washermen Area", 3)], "GOOD"),
    ],
    "Second Floor": [
        ("S3 Bathroom", [("Gyser", 1), ("Urinal", 1), ("Bath Area", 4), ("Latrin", 4)], "GOOD"),
        ("S4 Bathroom", [("Gyser", 1), ("Urinal", 2), ("Bath Area", 2), ("Latrin", 3), ("Washbasin", 3)], "GOOD"),
    ],
    "Third Floor": [
        ("T1 Bathroom", [("Latrin", 4), ("Bath Area", 4), ("Urinal", 1), ("Gyser", 1)], "GOOD"),
        ("T2 Bathroom", [("Latrin", 4), ("Bath Area", 2), ("Urinal", 1), ("Gyser", 1)], "GOOD"),
        ("T3A Bathroom", [("Latrin", 6)], "GOOD"),
        ("T3B Bath Area", [("Bath Area", 6), ("Gyser", 1), ("Washbasin", 1)], "GOOD"),
        ("T4 Bathroom", [("Urinal", 1), ("Bath Area", 5), ("Latrin", 6), ("Gyser", 1), ("Washbasin", 3)], "GOOD"),
        ("Third Floor Middle Common Area", [("Table Tennis", 1), ("Green Board", 1), ("Kitchen Basin", 1), ("Study Table", 2)], "GOOD"),
        ("Study Zone cum Balcony", [("Table", 1), ("Green Board", 1)], "GOOD"),
    ],
}


FLOOR_ROOMS: dict[str, list[RoomDef]] = {
    "Ground Floor": [
        room("101", "OFFICE", items=[("Table", 2), ("Chair", 4), ("Single Seater Sofa", 2), ("Three Seater Sofa", 1)]),
        room("102", "VISITOR_ROOM", "AC"),
        *[room(no, "STUDENT_ROOM", "AC", 3) for no in ["103", "104", "105", "106", "107", "108", "109", "110"]],
        room("111", "RESIDENCE"),
        room("112", "STORE"),
        *[room(no, "STUDENT_ROOM", "AC", 3) for no in ["113", "114", "115", "116", "117", "118", "119", "120", "121", "122"]],
    ],
    "First Floor": [
        room("Warden Office", "OFFICE"),
        room("201", "LOCKED_ROOM", "NON_AC", 0, "LOCKED"),
        room("202", "NON_AC_ROOM", "NON_AC", 0, "LOCKED"),
        room("203", "MERGED_NON_AC_ROOM", "NON_AC", 0, "MERGED", 1, "202"),
        room("204", "DAMAGED_ROOM", "AC", 0, "DAMAGED"),
        room("205", "MERGED_DAMAGED_ROOM", "AC", 0, "MERGED", 1, "204"),
        room("206", "NON_AC_ROOM", "NON_AC", 0, "LOCKED"),
        room("207", "MERGED_NON_AC_ROOM", "NON_AC", 0, "MERGED", 1, "206"),
        room("208", "NON_AC_ROOM", "NON_AC", 0, "LOCKED"),
        room("209", "MERGED_NON_AC_ROOM", "NON_AC", 0, "MERGED", 1, "208"),
        room("210", "DAMAGED_ROOM", "AC", 0, "DAMAGED"),
        room("211", "STUDENT_ROOM", "AC", 4),
        room("212", "STORE"),
        room("213", "NON_AC_ROOM", "NON_AC", 0, "LOCKED"),
        room("214", "NON_AC_ROOM", "NON_AC", 0, "LOCKED"),
        room("215", "NON_AC_ROOM", "NON_AC", 0, "LOCKED"),
        room("216", "NON_AC_ROOM", "NON_AC", 0, "LOCKED"),
        room("217", "NON_AC_ROOM", "NON_AC", 0, "LOCKED"),
        room("218", "DAMAGED_ROOM", "AC", 0, "DAMAGED"),
        room("219", "DAMAGED_ROOM", "AC", 0, "DAMAGED"),
        room("220", "NON_AC_ROOM", "NON_AC", 0, "LOCKED"),
        room("221", "STORE"),
        room("Study Room", "STUDY_ROOM", items=[("Big Table", 5), ("Small Table", 22), ("Chair", 50), ("TV", 1)]),
        room("222", "LOCKED_STUDENT_ROOM", "AC", 3, "LOCKED"),
        room("223", "STUDENT_ROOM", "AC", 3),
        room("224", "STUDENT_ROOM", "AC", 3),
        room("225", "STUDENT_ROOM", "AC", 3),
        room("226", "STUDENT_ROOM", "AC", 3),
        room("227", "DAMAGED_STUDENT_ROOM", "AC", 3, "DAMAGED"),
        room("228", "STUDENT_ROOM", "AC", 3),
        room("229", "STUDENT_ROOM", "AC", 3),
        room("230", "LAUNDRY_ROOM"),
        room("231", "STAFF_ROOM"),
        room("232", "DAMAGED_STUDENT_ROOM", "AC", 3, "DAMAGED"),
        room("233", "STUDENT_ROOM", "AC", 3),
        room("234", "LOCKED_ROOM", "NON_AC", 0, "LOCKED"),
        room("235", "LOCKED_ROOM", "NON_AC", 0, "LOCKED"),
        room("236", "LOCKED_ROOM", "NON_AC", 0, "LOCKED"),
        room("237", "STUDENT_ROOM", "AC", 3),
        room("238", "STUDENT_ROOM", "AC", 3),
        room("239", "DAMAGED_STUDENT_ROOM", "AC", 3, "DAMAGED"),
        room("240", "DAMAGED_STUDENT_ROOM", "AC", 3, "DAMAGED"),
        room("241", "DOCTOR_ROOM", "NON_AC", 0, "LOCKED"),
    ],
    "Second Floor": [
        *[room(str(no), "LOCKED_ROOM", "NON_AC", 0, "LOCKED") for no in range(301, 318)],
        room("318", "STUDENT_ROOM", "AC", 4),
        room("319", "LOCKED_STUDENT_ROOM", "AC", 3, "LOCKED"),
        room("320", "STUDENT_ROOM", "AC", 3),
        room("321", "DORMITORY", "AC", 5),
        room("322", "MERGED_DORMITORY", "AC", 0, "MERGED", 1, "321"),
        room("323", "DORMITORY", "AC", 5, items=[("Maid Bed", 1)]),
        room("324", "MERGED_DORMITORY", "AC", 0, "MERGED", 1, "323"),
        room("325", "WARDEN_RESIDENCE"),
        room("326", "STORE"),
        room("327", "STUDENT_ROOM", "AC", 3),
        room("328", "STUDENT_ROOM", "AC", 6),
        room("329", "LOCKED_STUDENT_ROOM", "AC", 5, "LOCKED"),
        room("330", "DORMITORY", "AC", 9, items=[("Maid Bed", 1)]),
        room("331", "MERGED_DORMITORY", "AC", 0, "MERGED", 1, "330"),
        room("332", "LOCKED_AC_ROOM", "AC", 0, "LOCKED"),
        room("Study Room", "STUDY_ROOM", items=[("Student Table", 10), ("Executive Table", 1), ("TV", 1), ("Chair", 49)]),
    ],
    "Third Floor": [
        room("401", "WARDEN_OFFICE"),
        *[room(str(no), "STUDENT_ROOM", "AC", 3) for no in range(402, 409)],
        room("409", "WARDEN_RESIDENCE"),
        *[room(str(no), "STUDENT_ROOM", "AC", 3) for no in range(410, 420)],
        room("420", "LOCKED_AC_ROOM", "AC", 0, "LOCKED"),
        room("421", "STUDENT_ROOM", "AC", 4),
        room("422", "STUDENT_ROOM", "AC", 4),
        room("423", "LOCKED_ROOM", "NON_AC", 0, "LOCKED"),
        room("424", "LOCKED_ROOM", "NON_AC", 0, "LOCKED"),
        room("425", "LOCKED_ROOM", "NON_AC", 0, "LOCKED"),
        room("426", "LOCKED_STUDENT_ROOM", "AC", 3, "LOCKED"),
        room("427", "STUDENT_ROOM", "AC", 3),
        room("428", "WARDEN_RESIDENCE"),
        room("429", "STUDENT_ROOM", "AC", 3),
        room("430", "STUDENT_ROOM", "AC", 3),
        room("431", "LOCKED_ROOM", "NON_AC", 0, "LOCKED"),
        room("432", "LOCKED_ROOM", "NON_AC", 0, "LOCKED"),
        room("433", "LOCKED_ROOM", "NON_AC", 0, "LOCKED"),
        room("434", "STORE", "NON_AC", 0, "LOCKED"),
        room("435", "LOCKED_ROOM", "NON_AC", 0, "LOCKED"),
        room("436", "STUDENT_ROOM", "AC", 3),
        room("437", "STUDENT_ROOM", "AC", 3),
        room("438", "LOCKED_AC_ROOM", "AC", 0, "LOCKED"),
        room("439", "STUDENT_ROOM", "AC", 3),
        room("440", "STUDENT_ROOM", "AC", 3),
        room("441", "STUDENT_ROOM", "AC", 3),
        room("442", "STUDENT_ROOM", "AC", 3),
        room("443", "STUDENT_ROOM", "AC", 3),
        room("444", "STUDENT_ROOM", "AC", 4),
        room("445", "LOCKED_ROOM", "NON_AC", 0, "LOCKED"),
        room("446", "LOCKED_ROOM", "NON_AC", 0, "LOCKED"),
    ],
}


ROOM_GROUPS = {
    "Ground-VI": ["103", "104", "105", "106", "107"],
    "Ground-VII": ["108", "109", "110", "113", "121", "122", "114"],
    "Ground-VIII": ["114", "115", "116", "117", "118", "119", "120"],
    "First-XI": ["211"],
    "First-XII": ["223", "224", "225"],
    "Second-I": ["318"],
    "Second-II": ["320"],
    "Second-III": ["327"],
    "Second-IV": ["321"],
    "Second-V": ["328"],
    "Third-IX": ["402", "403", "404", "405", "406", "407", "408", "410", "411"],
    "Third-X": ["412", "413", "414", "415", "416", "417", "418", "419", "421"],
}


def parse_students() -> list[dict[str, str]]:
    if not STUDENTS_XLSX.exists():
        raise FileNotFoundError(f"Student file not found: {STUDENTS_XLSX}")

    wb = load_workbook(STUDENTS_XLSX, read_only=True, data_only=True)
    ws = wb["Sheet1"]
    students: list[dict[str, str]] = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[1] or not row[2] or not row[3]:
            continue
        class_section = str(row[2]).strip().split()
        class_name = class_section[0]
        section = class_section[1] if len(class_section) > 1 else ""
        students.append(
            {
                "ledger_no": str(row[1]).strip(),
                "student_name": str(row[3]).strip(),
                "class": class_name,
                "section": section,
            }
        )

    wb.close()
    return students


def room_sort_key(room_no: str) -> tuple[int, str]:
    digits = "".join(ch for ch in room_no if ch.isdigit())
    return (int(digits) if digits else 9999, room_no)


def build_room_plan(students: list[dict[str, str]]) -> dict[str, list[dict[str, str]]]:
    grouped: dict[str, list[dict[str, str]]] = defaultdict(list)
    for student in students:
        grouped[student["class"]].append(student)

    for cls in grouped:
        grouped[cls].sort(key=lambda s: (s["section"], s["student_name"], s["ledger_no"]))

    assignments = {
        "103": grouped["VI"][:3],
        "104": grouped["VI"][3:6],
        "105": grouped["VI"][6:9],
        "106": grouped["VI"][9:12],
        "107": grouped["VI"][12:13],
        "108": grouped["VII"][:3],
        "109": grouped["VII"][3:6],
        "110": grouped["VII"][6:9],
        "113": grouped["VII"][9:12],
        "121": grouped["VII"][12:15],
        "122": grouped["VII"][15:18],
        "114": grouped["VII"][18:19] + grouped["VIII"][:2],
        "115": grouped["VIII"][2:5],
        "116": grouped["VIII"][5:8],
        "117": grouped["VIII"][8:11],
        "118": grouped["VIII"][11:14],
        "119": grouped["VIII"][14:17],
        "120": grouped["VIII"][17:20],
        "211": grouped["XI"],
        "223": grouped["XII"][:3],
        "224": grouped["XII"][3:6],
        "225": grouped["XII"][6:7],
        "318": grouped["I"],
        "320": grouped["II"],
        "327": grouped["III"],
        "321": grouped["IV"],
        "328": grouped["V"],
        "402": grouped["IX"][:3],
        "403": grouped["IX"][3:6],
        "404": grouped["IX"][6:9],
        "405": grouped["IX"][9:12],
        "406": grouped["IX"][12:15],
        "407": grouped["IX"][15:18],
        "408": grouped["IX"][18:21],
        "410": grouped["IX"][21:24],
        "411": grouped["IX"][24:25],
        "412": grouped["X"][:3],
        "413": grouped["X"][3:6],
        "414": grouped["X"][6:9],
        "415": grouped["X"][9:12],
        "416": grouped["X"][12:15],
        "417": grouped["X"][15:18],
        "418": grouped["X"][18:21],
        "419": grouped["X"][21:24],
        "421": grouped["X"][24:26],
    }
    return assignments


def main() -> None:
    students = parse_students()
    room_plan = build_room_plan(students)

    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()
    cur.execute("PRAGMA foreign_keys = OFF")

    # Remove current structure + student data but preserve auth/settings data.
    delete_order = [
        "bed_allotments",
        "beds",
        "room_items",
        "common_area_items",
        "common_areas",
        "rooms",
        "floors",
        "blocks",
        "hostels",
        "students",
        "student_biodata",
        "student_parents",
        "student_guardians",
        "student_profile",
        "student_behaviour_log",
        "student_permissions",
        "student_permission_files",
        "student_self_going",
        "student_device_permission",
        "student_pocket_wallet",
        "expected_rooms",
    ]
    for table in delete_order:
        cur.execute(f'DELETE FROM "{table}"')

    cur.execute("INSERT INTO hostels (name) VALUES (?)", (HOSTEL_NAME,))
    hostel_id = cur.lastrowid

    cur.execute("INSERT INTO blocks (hostel_id, name) VALUES (?, ?)", (hostel_id, BLOCK_NAME))
    block_id = cur.lastrowid

    floor_ids: dict[str, int] = {}
    for floor_name in ["Ground Floor", "First Floor", "Second Floor", "Third Floor"]:
        cur.execute("INSERT INTO floors (school_id, floor_name) VALUES (?, ?)", (hostel_id, floor_name))
        floor_ids[floor_name] = cur.lastrowid

    room_ids: dict[str, int] = {}
    merged_links: list[tuple[int, str]] = []
    bed_ids_by_room: dict[str, list[int]] = defaultdict(list)

    for floor_name, room_defs in FLOOR_ROOMS.items():
        for definition in room_defs:
            cur.execute(
                """
                INSERT INTO rooms
                (floor_id, room_no, status, room_type, bed_capacity, ac_type, is_merged, merged_into_room_id)
                VALUES (?, ?, ?, ?, ?, ?, ?, NULL)
                """,
                (
                    floor_ids[floor_name],
                    definition.room_no,
                    definition.status,
                    definition.room_type,
                    definition.bed_capacity,
                    definition.ac_type,
                    definition.is_merged,
                ),
            )
            room_id = cur.lastrowid
            room_ids[definition.room_no] = room_id

            if definition.merged_into_room_no:
                merged_links.append((room_id, definition.merged_into_room_no))

            for item_name, qty in definition.items:
                cur.execute(
                    """
                    INSERT INTO room_items (room_id, item_name, quantity, status, last_updated)
                    VALUES (?, ?, ?, 'PRESENT', ?)
                    """,
                    (room_id, item_name, qty, TODAY),
                )

            for bed_index in range(1, definition.bed_capacity + 1):
                cur.execute(
                    "INSERT INTO beds (room_id, bed_no, status) VALUES (?, ?, 'VACANT')",
                    (room_id, str(bed_index)),
                )
                bed_ids_by_room[definition.room_no].append(cur.lastrowid)

            cur.execute(
                "INSERT INTO expected_rooms (floor_no, room_no, room_type, status) VALUES (?, ?, ?, ?)",
                (floor_name, definition.room_no, definition.room_type, definition.status),
            )

    for room_id, merged_room_no in merged_links:
        cur.execute(
            "UPDATE rooms SET merged_into_room_id = ? WHERE id = ?",
            (room_ids[merged_room_no], room_id),
        )

    for floor_name, areas in FLOOR_COMMON_AREAS.items():
        for area_name, items, status in areas:
            cur.execute(
                "INSERT INTO common_areas (floor_id, area_name, status) VALUES (?, ?, ?)",
                (floor_ids[floor_name], area_name, status),
            )
            area_id = cur.lastrowid
            for item_name, qty in items:
                cur.execute(
                    """
                    INSERT INTO common_area_items
                    (common_area_id, item_name, quantity, status, remarks, last_updated)
                    VALUES (?, ?, ?, 'PRESENT', '', ?)
                    """,
                    (area_id, item_name, qty, TODAY),
                )

    session_row = cur.execute(
        "SELECT id FROM academic_sessions WHERE is_active = 1 ORDER BY id DESC LIMIT 1"
    ).fetchone()
    session_id = session_row["id"] if session_row else None

    student_ids: dict[str, int] = {}
    for student in students:
        cur.execute(
            """
            INSERT INTO students
            (ledger_no, student_name, class, section, session_id, is_active, created_at, biodata_completed, status)
            VALUES (?, ?, ?, ?, ?, 1, ?, 1, 'ACTIVE')
            """,
            (
                student["ledger_no"],
                student["student_name"],
                student["class"],
                student["section"],
                session_id,
                NOW,
            ),
        )
        student_id = cur.lastrowid
        student_ids[student["ledger_no"]] = student_id
        cur.execute(
            """
            INSERT INTO student_parents
            (student_id, father_name, father_mobile_1, father_mobile_2, mother_name, mother_mobile_1, mother_mobile_2)
            VALUES (?, '', '', '', '', '', '')
            """,
            (student_id,),
        )
        cur.execute(
            """
            INSERT INTO student_profile
            (ledger_no, admission_no, full_name, class, section, joining_date, status, remarks, created_at)
            VALUES (?, ?, ?, ?, ?, ?, 'ACTIVE', 'Imported from ASPCS hostel list', ?)
            """,
            (
                student["ledger_no"],
                student["ledger_no"],
                student["student_name"],
                student["class"],
                student["section"],
                TODAY,
                NOW,
            ),
        )
        cur.execute(
            """
            INSERT INTO student_pocket_wallet
            (ledger_no, opening_balance, current_balance, status, created_at)
            VALUES (?, 0, 0, 'ACTIVE', ?)
            """,
            (student["ledger_no"], NOW),
        )

    allotted_count = 0
    for room_no in sorted(room_plan, key=room_sort_key):
        occupants = room_plan[room_no]
        if not occupants:
            continue
        beds = bed_ids_by_room[room_no]
        for bed_id, student in zip(beds, occupants):
            cur.execute(
                """
                INSERT INTO bed_allotments
                (ledger_number, bed_id, allot_date, status, created_at)
                VALUES (?, ?, ?, 'ALLOTTED', ?)
                """,
                (student["ledger_no"], bed_id, TODAY, NOW),
            )
            cur.execute("UPDATE beds SET status = 'OCCUPIED' WHERE id = ?", (bed_id,))
            cur.execute(
                """
                INSERT INTO student_biodata
                (student_id, room_no, class_of_admission, hostel_reporting_date, remarks)
                VALUES (?, ?, ?, ?, ?)
                """,
                (
                    student_ids[student["ledger_no"]],
                    room_no,
                    student["class"],
                    TODAY,
                    "Auto-allotted during ASPCS hostel import",
                ),
            )
            allotted_count += 1

    conn.commit()

    summary = {
        "hostels": cur.execute("SELECT COUNT(*) FROM hostels").fetchone()[0],
        "blocks": cur.execute("SELECT COUNT(*) FROM blocks").fetchone()[0],
        "floors": cur.execute("SELECT COUNT(*) FROM floors").fetchone()[0],
        "rooms": cur.execute("SELECT COUNT(*) FROM rooms").fetchone()[0],
        "beds": cur.execute("SELECT COUNT(*) FROM beds").fetchone()[0],
        "common_areas": cur.execute("SELECT COUNT(*) FROM common_areas").fetchone()[0],
        "common_area_items": cur.execute("SELECT COUNT(*) FROM common_area_items").fetchone()[0],
        "room_items": cur.execute("SELECT COUNT(*) FROM room_items").fetchone()[0],
        "students": cur.execute("SELECT COUNT(*) FROM students").fetchone()[0],
        "bed_allotments": cur.execute("SELECT COUNT(*) FROM bed_allotments WHERE status = 'ALLOTTED'").fetchone()[0],
    }
    conn.close()

    print("ASPCS hostel setup import completed.")
    for key, value in summary.items():
        print(f"{key}\t{value}")
    print(f"students_allotted\t{allotted_count}")


if __name__ == "__main__":
    main()
