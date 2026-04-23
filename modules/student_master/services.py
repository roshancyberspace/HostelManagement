import base64
import os
import shutil
from datetime import datetime
from werkzeug.utils import secure_filename
from flask import current_app
from openpyxl import Workbook

from models.db import get_db
from models.floor_model import floor_label_sql
from modules.bed_allotment.service import auto_vacate_bed_on_status_change


# --------------------------------------------------
# ACTIVE ACADEMIC SESSION
# --------------------------------------------------
def get_active_session():
    db = get_db()
    cur = db.cursor()
    cur.execute("""
        SELECT * FROM academic_sessions
        WHERE is_active = 1 AND is_deleted = 0
        LIMIT 1
    """)
    row = cur.fetchone()
    db.close()
    return row


# --------------------------------------------------
# BASIC BULK ADD (LEDGER / NAME / CLASS)
# --------------------------------------------------
def bulk_add_students(rows):
    session = get_active_session()
    if not session:
        raise Exception("No active academic session")

    db = get_db()
    cur = db.cursor()

    for r in rows:
        if not r.get("ledger_no"):
            continue

        cur.execute("""
            INSERT OR IGNORE INTO students
            (ledger_no, student_name, class, section,
             session_id, status, biodata_completed)
            VALUES (?, ?, ?, ?, ?, 'ACTIVE', 0)
        """, (
            r["ledger_no"],
            r.get("name"),
            r.get("class"),
            r.get("section"),
            session["id"]
        ))

    db.commit()
    db.close()


# --------------------------------------------------
# FULL STUDENT FETCH (LEDGER BASED)
# --------------------------------------------------
def get_student_full_by_ledger(ledger):
    db = get_db()
    cur = db.cursor()

    cur.execute("SELECT * FROM students WHERE ledger_no=?", (ledger,))
    student = cur.fetchone()
    if not student:
        db.close()
        return None

    cur.execute("SELECT * FROM student_biodata WHERE student_id=?", (student["id"],))
    biodata = cur.fetchone()

    cur.execute("SELECT * FROM student_parents WHERE student_id=?", (student["id"],))
    parents = cur.fetchone()

    cur.execute("SELECT * FROM student_guardians WHERE student_id=?", (student["id"],))
    guardians = cur.fetchall()

    db.close()
    return {
        "student": student,
        "biodata": biodata,
        "parents": parents,
        "guardians": guardians
    }


# --------------------------------------------------
# PHOTO SAVE
# --------------------------------------------------
def get_student_upload_folder(ledger):
    base_folder = os.path.join(
        current_app.instance_path,
        "student_uploads", "students"
    )
    student_folder = os.path.join(base_folder, ledger)

    try:
        os.makedirs(student_folder, exist_ok=True)
        return student_folder, f"instance_uploads/students/{ledger}"
    except PermissionError:
        os.makedirs(base_folder, exist_ok=True)
        return base_folder, "instance_uploads/students"


def save_photo(file, ledger, name):
    if not file or not file.filename:
        return None

    folder, public_folder = get_student_upload_folder(ledger)
    ext = os.path.splitext(file.filename)[1]
    if public_folder.endswith(f"/{ledger}"):
        fname = secure_filename(f"{name}{ext}")
    else:
        fname = secure_filename(f"{ledger}_{name}{ext}")
    path = os.path.join(folder, fname)
    file.save(path)

    return f"{public_folder}/{fname}"


def save_base64_photo(data_url, ledger, name):
    if not data_url or "," not in data_url:
        return None

    try:
        header, encoded = data_url.split(",", 1)
        if "image/jpeg" in header:
            ext = ".jpg"
        elif "image/png" in header:
            ext = ".png"
        else:
            ext = ".jpg"

        raw = base64.b64decode(encoded)
    except Exception:
        return None

    folder, public_folder = get_student_upload_folder(ledger)
    if public_folder.endswith(f"/{ledger}"):
        fname = secure_filename(f"{name}{ext}")
    else:
        fname = secure_filename(f"{ledger}_{name}{ext}")
    path = os.path.join(folder, fname)
    with open(path, "wb") as image_file:
        image_file.write(raw)

    return f"{public_folder}/{fname}"


# --------------------------------------------------
# SAVE / UPDATE BIODATA
# --------------------------------------------------
def save_biodata(student_id, ledger, form, files):
    db = get_db()
    cur = db.cursor()

    student_photo = save_photo(files.get("student_photo"), ledger, "student")
    if not student_photo:
        student_photo = save_base64_photo(
            form.get("captured_student_photo"),
            ledger,
            "student"
        )
    father_photo = save_photo(files.get("father_photo"), ledger, "father")
    mother_photo = save_photo(files.get("mother_photo"), ledger, "mother")

    cur.execute(
        "SELECT id FROM student_biodata WHERE student_id = ?",
        (student_id,)
    )
    biodata_exists = cur.fetchone()

    if biodata_exists:
        cur.execute("""
            UPDATE student_biodata
            SET dob = ?,
                class_of_admission = ?,
                date_of_admission = ?,
                hostel_reporting_date = ?,
                permanent_address = ?,
                remarks = ?,
                student_photo = COALESCE(?, student_photo)
            WHERE student_id = ?
        """, (
            form.get("dob"),
            form.get("class_of_admission"),
            form.get("date_of_admission"),
            form.get("hostel_reporting_date"),
            form.get("permanent_address"),
            form.get("remarks"),
            student_photo,
            student_id
        ))
    else:
        cur.execute("""
            INSERT INTO student_biodata
            (student_id, dob, class_of_admission, date_of_admission,
             hostel_reporting_date, permanent_address, remarks, student_photo)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
        """, (
            student_id,
            form.get("dob"),
            form.get("class_of_admission"),
            form.get("date_of_admission"),
            form.get("hostel_reporting_date"),
            form.get("permanent_address"),
            form.get("remarks"),
            student_photo
        ))

    cur.execute(
        "SELECT student_id FROM student_parents WHERE student_id = ?",
        (student_id,)
    )
    parents_exists = cur.fetchone()

    if parents_exists:
        cur.execute("""
            UPDATE student_parents
            SET father_name = ?,
                father_mobile_1 = ?,
                father_photo = COALESCE(?, father_photo),
                mother_name = ?,
                mother_mobile_1 = ?,
                mother_photo = COALESCE(?, mother_photo)
            WHERE student_id = ?
        """, (
            form.get("father_name"),
            form.get("father_mobile_1"),
            father_photo,
            form.get("mother_name"),
            form.get("mother_mobile_1"),
            mother_photo,
            student_id
        ))
    else:
        cur.execute("""
            INSERT INTO student_parents
            (student_id, father_name, father_mobile_1, father_photo,
             mother_name, mother_mobile_1, mother_photo)
            VALUES (?, ?, ?, ?, ?, ?, ?)
        """, (
            student_id,
            form.get("father_name"),
            form.get("father_mobile_1"),
            father_photo,
            form.get("mother_name"),
            form.get("mother_mobile_1"),
            mother_photo
        ))

    cur.execute("DELETE FROM student_guardians WHERE student_id=?", (student_id,))
    names = form.getlist("guardian_name[]")
    relations = form.getlist("guardian_relation[]")
    mobiles = form.getlist("guardian_mobile[]")
    addresses = form.getlist("guardian_address[]")
    photos = files.getlist("guardian_photo[]")

    for i in range(len(names)):
        if names[i] and relations[i] and mobiles[i]:
            gphoto = save_photo(photos[i], ledger, f"guardian_{i+1}")
            cur.execute("""
                INSERT INTO student_guardians
                (student_id, name, relation, mobile_1, address, photo)
                VALUES (?, ?, ?, ?, ?, ?)
            """, (
                student_id,
                names[i],
                relations[i],
                mobiles[i],
                addresses[i],
                gphoto
            ))

    cur.execute("""
        UPDATE students
        SET biodata_completed = 1
        WHERE id = ?
    """, (student_id,))

    db.commit()
    db.close()


# --------------------------------------------------
# EXCEL TEMPLATES
# --------------------------------------------------
def generate_basic_student_excel():
    wb = Workbook()
    ws = wb.active
    ws.title = "BASIC_STUDENT_UPLOAD"
    ws.append(["Ledger Number*", "Student Name*", "Class*", "Section"])
    ws.append(["L12345", "Ravi Kumar", "VI", "A"])
    return wb


def generate_full_biodata_excel():
    wb = Workbook()
    ws = wb.active
    ws.title = "STUDENT_DETAILS"
    ws.append([
        "Ledger Number*", "Student Name", "Class", "Section",
        "DOB", "Class of Admission", "Date of Admission",
        "Hostel Reporting Date", "Permanent Address", "Remarks"
    ])
    wb.create_sheet("PARENTS_DETAILS")
    wb.create_sheet("GUARDIANS_DETAILS")
    wb.create_sheet("INSTRUCTIONS")
    return wb


# --------------------------------------------------
# CENTRAL STUDENT STATUS CHANGE (FINAL & FROZEN)
# --------------------------------------------------
def change_student_status(ledger_no, new_status, reason=None):
    """
    SINGLE SOURCE OF TRUTH for student status changes

    - Updates student status
    - Auto-vacates bed (ledger-based, history safe)
    - Revokes all permissions
    """

    db = get_db()
    cur = db.cursor()

    # Validate student
    cur.execute(
        "SELECT id FROM students WHERE ledger_no = ?",
        (ledger_no,)
    )
    student = cur.fetchone()
    if not student:
        db.close()
        raise Exception("Student not found")

    student_id = student["id"]

    # 1️⃣ Update status
    cur.execute("""
        UPDATE students
        SET status = ?,
            status_updated_at = CURRENT_TIMESTAMP
        WHERE id = ?
    """, (new_status, student_id))

    # 2️⃣ AUTO BED VACATE (via bed_allotments)
    auto_vacate_bed_on_status_change(cur, ledger_no, new_status)

    # 3️⃣ Revoke self-going permissions
    cur.execute("""
        UPDATE student_self_going
        SET allowed = 0,
            revoked_at = ?,
            revoked_reason = ?
        WHERE student_id = ?
          AND allowed = 1
    """, (datetime.now(), reason, student_id))

    # 4️⃣ Revoke device permissions
    cur.execute("""
        UPDATE student_device_permission
        SET allowed = 0,
            revoked_at = ?,
            revoked_reason = ?
        WHERE student_id = ?
          AND allowed = 1
    """, (datetime.now(), reason, student_id))

    db.commit()
    db.close()


def delete_student_record(ledger_no):
    ledger_no = (ledger_no or "").strip()
    if not ledger_no:
        raise Exception("Ledger number is required")

    db = get_db()
    cur = db.cursor()

    cur.execute(
        "SELECT id FROM students WHERE ledger_no = ?",
        (ledger_no,)
    )
    student = cur.fetchone()
    if not student:
        db.close()
        raise Exception("Student not found")

    student_id = student["id"]

    cur.execute("""
        DELETE FROM student_permission_files
        WHERE permission_id IN (
            SELECT id FROM student_permissions WHERE student_id = ?
        )
    """, (student_id,))

    for table in (
        "student_permissions",
        "student_self_going",
        "student_device_permission",
        "student_guardians",
        "student_parents",
        "student_biodata",
    ):
        cur.execute(f"DELETE FROM {table} WHERE student_id = ?", (student_id,))

    for table, column in (
        ("bed_allotments", "ledger_number"),
        ("barber_haircut_records", "ledger_no"),
        ("gate_passes", "ledger_no"),
        ("hostel_decisions", "ledger_no"),
        ("hostel_fee_payments", "ledger_no"),
        ("hostel_notice_ack", "ledger_no"),
        ("laundry_register", "ledger_no"),
        ("pocket_money_transactions", "ledger_no"),
        ("student_attendance", "ledger_no"),
        ("student_behaviour_log", "ledger_no"),
        ("student_pocket_wallet", "ledger_no"),
        ("student_profile", "ledger_no"),
    ):
        cur.execute(f"DELETE FROM {table} WHERE {column} = ?", (ledger_no,))

    cur.execute("DELETE FROM students WHERE id = ?", (student_id,))
    db.commit()
    db.close()

    upload_root = os.path.join(
        current_app.instance_path,
        "student_uploads",
        "students",
        ledger_no
    )
    shutil.rmtree(upload_root, ignore_errors=True)


# --------------------------------------------------
# BED + PERMISSION READ-ONLY HELPERS
# --------------------------------------------------

# ✅ FIXED: Now uses beds -> rooms -> floors (NOT beds_master)
def get_student_bed_info(ledger_no):
    db = get_db()
    cur = db.cursor()

    # Current bed
    cur.execute(f"""
        SELECT
            {floor_label_sql("f")} AS floor_no,
            r.room_no,
            b.bed_no,
            a.allot_date
        FROM bed_allotments a
        JOIN beds b ON a.bed_id = b.id
        JOIN rooms r ON b.room_id = r.id
        JOIN floors f ON r.floor_id = f.id
        WHERE a.ledger_number = ?
          AND a.status = 'ALLOTTED'
        ORDER BY a.id DESC
        LIMIT 1
    """, (ledger_no,))
    current_bed = cur.fetchone()

    # Bed history
    cur.execute(f"""
        SELECT
            {floor_label_sql("f")} AS floor_no,
            r.room_no,
            b.bed_no,
            a.allot_date,
            a.vacate_date,
            a.vacate_reason
        FROM bed_allotments a
        JOIN beds b ON a.bed_id = b.id
        JOIN rooms r ON b.room_id = r.id
        JOIN floors f ON r.floor_id = f.id
        WHERE a.ledger_number = ?
        ORDER BY a.id DESC
    """, (ledger_no,))
    history = cur.fetchall()

    db.close()
    return current_bed, history


# ✅ FIXED: Removed from_date/to_date (your DB doesn't have these)
def get_student_permissions(student_id):
    db = get_db()
    cur = db.cursor()

    # Self-going history (SAFE)
    cur.execute("""
        SELECT allowed,
               revoked_at, revoked_reason,
               created_at
        FROM student_self_going
        WHERE student_id = ?
        ORDER BY id DESC
    """, (student_id,))
    self_going = cur.fetchall()

    # Device permission history (SAFE)
    cur.execute("""
        SELECT device_type, allowed,
               revoked_at, revoked_reason,
               created_at
        FROM student_device_permission
        WHERE student_id = ?
        ORDER BY id DESC
    """, (student_id,))
    devices = cur.fetchall()

    db.close()
    return self_going, devices


# --------------------------------------------------
# BASIC STUDENT BULK EXCEL (LEDGER / NAME / CLASS)
# --------------------------------------------------
def generate_basic_student_bulk_excel():
    wb = Workbook()
    ws = wb.active
    ws.title = "STUDENT_BULK_UPLOAD"

    ws.append(["Ledger Number*", "Student Name*", "Class*"])
    ws.append(["L1001", "Amit Kumar", "VI"])
    ws.append(["L1002", "Rohit Singh", "VII"])

    return wb
