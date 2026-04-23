from openpyxl import load_workbook
from models.db import get_db
from .services import get_active_session, bulk_add_students


# ==================================================
# BASIC STUDENT BULK EXCEL (LEDGER / NAME / CLASS)
# ==================================================
def process_basic_student_excel(filepath):
    """
    Excel columns:
    Ledger Number | Student Name | Class

    Purpose:
    - Build initial student registry
    - Status = ACTIVE
    - Biodata filled later
    """

    wb = load_workbook(filepath)
    ws = wb.active

    rows = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        ledger_no, student_name, class_name = row

        if not ledger_no:
            continue

        rows.append({
            "ledger_no": str(ledger_no).strip(),
            "name": str(student_name).strip() if student_name else None,
            "class": str(class_name).strip() if class_name else None,
            "section": None
        })

    if rows:
        bulk_add_students(rows)


# ==================================================
# FULL BIODATA EXCEL (AUTHORITATIVE – FINAL)
# ==================================================
def process_biodata_excel(filepath):
    """
    Process FULL BIODATA Excel and update database.

    Expected sheets:
    - STUDENT_DETAILS
    - PARENTS_DETAILS
    - GUARDIANS_DETAILS
    """

    wb = load_workbook(filepath)

    session = get_active_session()
    if not session:
        raise Exception("No active academic session")

    db = get_db()
    cur = db.cursor()

    # --------------------------------------------------
    # STUDENT DETAILS
    # --------------------------------------------------
    ws = wb["STUDENT_DETAILS"]
    for row in ws.iter_rows(min_row=2, values_only=True):
        (
            ledger_no, student_name, class_name, section,
            dob, class_of_admission, date_of_admission,
            hostel_reporting_date, permanent_address,
            remarks
        ) = row

        if not ledger_no:
            continue

        # Fetch or create student
        cur.execute("SELECT id FROM students WHERE ledger_no=?", (ledger_no,))
        s = cur.fetchone()

        if not s:
            cur.execute("""
                INSERT INTO students
                (ledger_no, student_name, class, section,
                 session_id, status, biodata_completed)
                VALUES (?, ?, ?, ?, ?, 'ACTIVE', 0)
            """, (
                ledger_no, student_name, class_name,
                section, session["id"]
            ))
            student_id = cur.lastrowid
        else:
            student_id = s["id"]
            cur.execute("""
                UPDATE students
                SET student_name=?, class=?, section=?
                WHERE id=?
            """, (student_name, class_name, section, student_id))

        # Biodata upsert
        cur.execute("""
            INSERT INTO student_biodata
            (student_id, dob, class_of_admission,
             date_of_admission, hostel_reporting_date,
             permanent_address, remarks)
            VALUES (?, ?, ?, ?, ?, ?, ?)
            ON CONFLICT(student_id) DO UPDATE SET
                dob=excluded.dob,
                class_of_admission=excluded.class_of_admission,
                date_of_admission=excluded.date_of_admission,
                hostel_reporting_date=excluded.hostel_reporting_date,
                permanent_address=excluded.permanent_address,
                remarks=excluded.remarks
        """, (
            student_id, dob, class_of_admission,
            date_of_admission, hostel_reporting_date,
            permanent_address, remarks
        ))

    # --------------------------------------------------
    # PARENTS DETAILS
    # --------------------------------------------------
    ws = wb["PARENTS_DETAILS"]
    for row in ws.iter_rows(min_row=2, values_only=True):
        ledger_no, father_name, father_mobile, mother_name, mother_mobile = row

        if not ledger_no:
            continue

        cur.execute("SELECT id FROM students WHERE ledger_no=?", (ledger_no,))
        s = cur.fetchone()
        if not s:
            continue

        cur.execute("""
            INSERT INTO student_parents
            (student_id, father_name, father_mobile_1,
             mother_name, mother_mobile_1)
            VALUES (?, ?, ?, ?, ?)
            ON CONFLICT(student_id) DO UPDATE SET
                father_name=excluded.father_name,
                father_mobile_1=excluded.father_mobile_1,
                mother_name=excluded.mother_name,
                mother_mobile_1=excluded.mother_mobile_1
        """, (
            s["id"], father_name, father_mobile,
            mother_name, mother_mobile
        ))

    # --------------------------------------------------
    # GUARDIANS DETAILS
    # --------------------------------------------------
    ws = wb["GUARDIANS_DETAILS"]
    for row in ws.iter_rows(min_row=2, values_only=True):
        ledger_no, g_name, g_relation, g_mobile, g_address = row

        if not ledger_no or not g_name or not g_relation or not g_mobile:
            continue

        cur.execute("SELECT id FROM students WHERE ledger_no=?", (ledger_no,))
        s = cur.fetchone()
        if not s:
            continue

        cur.execute("""
            INSERT INTO student_guardians
            (student_id, name, relation, mobile_1, address)
            VALUES (?, ?, ?, ?, ?)
        """, (
            s["id"], g_name, g_relation, g_mobile, g_address
        ))

    # --------------------------------------------------
    # FINALIZE
    # --------------------------------------------------
    cur.execute("""
        UPDATE students
        SET biodata_completed = 1
        WHERE id IN (SELECT DISTINCT student_id FROM student_biodata)
    """)

    db.commit()
    db.close()
